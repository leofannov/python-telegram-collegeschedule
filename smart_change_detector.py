import hashlib
import json
import os
from datetime import datetime
from openpyxl import load_workbook
from config import RANGES, EXCEL_FILE
import difflib

class SmartChangeDetector:
    def __init__(self, cache_file='cache/smart_schedule_cache.json'):
        self.cache_file = cache_file
        
        if not os.path.exists('cache'):
            os.makedirs('cache')

    def extract_schedule_data(self, group: str):
        """Извлечь данные расписания - ТОЛЬКО schedule диапазоны"""
        try:
            if not os.path.exists(EXCEL_FILE):
                print(f"❌ Файл {EXCEL_FILE} не существует")
                return None

            if group not in RANGES:
                print(f"❌ Группа {group} не найдена в RANGES")
                return None

            print(f"📖 Чтение данных для группы {group}...")
            wb = load_workbook(EXCEL_FILE, data_only=True)
            ws = wb.active
            
            schedule_data = {
                'group': group,
                'weeks': {},
                'extracted_at': datetime.now().isoformat()
            }
            
            for week_type in ['even', 'odd']:
                if week_type not in RANGES[group]:
                    continue
                    
                schedule_data['weeks'][week_type] = {}
                
                for day, day_ranges in RANGES[group][week_type].items():
                    print(f"  📅 Обработка {day} ({week_type} неделя)")
                    
                    day_data = {
                        'schedule': []  # ТОЛЬКО schedule данные
                    }
                    
                    # Читаем ТОЛЬКО schedule диапазон
                    if 'schedule' in day_ranges:
                        cell_range = day_ranges['schedule']
                        try:
                            cells = ws[cell_range]
                            range_data = []
                            for row in cells:
                                row_data = []
                                for cell in row:
                                    value = cell.value
                                    clean_value = str(value).strip() if value is not None else ""
                                    row_data.append(clean_value)
                                # Сохраняем даже пустые строки для точного сравнения
                                range_data.append(row_data)
                            day_data['schedule'] = range_data
                            print(f"    📊 schedule ({cell_range}): {len(range_data)} строк")
                            
                            # Показать содержимое для отладки
                            for i, row in enumerate(range_data):
                                if any(cell.strip() for cell in row):
                                    print(f"      Строка {i+1}: {row}")
                                    
                        except Exception as e:
                            print(f"    ❌ Ошибка чтения {cell_range}: {e}")
                            day_data['schedule'] = []
                    
                    schedule_data['weeks'][week_type][day] = day_data
            
            wb.close()
            print(f"✅ Данные успешно извлечены для {group}")
            return schedule_data
            
        except Exception as e:
            print(f"❌ Критическая ошибка извлечения данных для {group}: {e}")
            return None

    def calculate_smart_hash(self, group: str):
        """Умный расчет хэша - ТОЛЬКО schedule данные"""
        schedule_data = self.extract_schedule_data(group)
        if not schedule_data:
            return None
        
        # Создаем строку для хэширования, исключая метаданные
        hash_data = {
            'weeks': schedule_data['weeks']
        }
        
        data_string = json.dumps(hash_data, sort_keys=True, ensure_ascii=False)
        file_hash = hashlib.md5(data_string.encode('utf-8')).hexdigest()
        
        print(f"🔍 Умный хэш для {group}: {file_hash[:16]}...")
        return file_hash

    def has_changed(self, group: str):
        """Проверка изменений - ТОЛЬКО schedule данные"""
        print(f"\n🎯 ПРОВЕРКА ИЗМЕНЕНИЙ: {group}")
        
        current_hash = self.calculate_smart_hash(group)
        if not current_hash:
            print("❌ Не удалось вычислить хэш")
            return False, "Не удалось вычислить хэш"

        old_data = self.get_old_data(group)
        
        if not old_data or 'hash' not in old_data:
            print(f"📝 Первый запуск для {group}, сохраняем данные")
            self.save_schedule_data(group, self.extract_schedule_data(group), current_hash)
            return False, "Первый запуск"
        
        old_hash = old_data.get('hash')
        
        if current_hash == old_hash:
            print(f"✅ Изменений нет для {group}")
            return False, "Хэши совпадают"
        
        print(f"🔄 ИЗМЕНЕНИЯ ОБНАРУЖЕНЫ для {group}!")
        print(f"   Старый хэш: {old_hash[:16]}...")
        print(f"   Новый хэш:  {current_hash[:16]}...")
        
        # Детальный анализ изменений
        current_data = self.extract_schedule_data(group)
        changes = self.analyze_changes(group, old_data, current_data)
        
        # Сохраняем новые данные
        self.save_schedule_data(group, current_data, current_hash)
        
        return True, changes

    def analyze_changes(self, group: str, old_data: dict, new_data: dict):
        """Анализ конкретных изменений - ТОЛЬКО schedule"""
        changes = []
        
        try:
            for week_type in ['even', 'odd']:
                if week_type in old_data.get('weeks', {}) and week_type in new_data.get('weeks', {}):
                    for day in ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']:
                        if day in old_data['weeks'][week_type] and day in new_data['weeks'][week_type]:
                            old_day_data = old_data['weeks'][week_type][day]
                            new_day_data = new_data['weeks'][week_type][day]
                            
                            # Сравниваем ТОЛЬКО schedule данные
                            old_schedule = old_day_data.get('schedule', [])
                            new_schedule = new_day_data.get('schedule', [])
                            
                            if old_schedule != new_schedule:
                                change_desc = f"{week_type}_{day}"
                                changes.append(change_desc)
                                print(f"   📝 Изменения в {day} ({week_type} неделя)")
                                
                                # Детальное сравнение значений
                                max_rows = max(len(old_schedule), len(new_schedule))
                                for i in range(max_rows):
                                    old_row = old_schedule[i] if i < len(old_schedule) else []
                                    new_row = new_schedule[i] if i < len(new_schedule) else []
                                    
                                    if old_row != new_row:
                                        print(f"      Строка {i+1}:")
                                        print(f"        Старое: {old_row}")
                                        print(f"        Новое:  {new_row}")
        
        except Exception as e:
            print(f"❌ Ошибка анализа изменений: {e}")
        
        return changes

    def get_old_data(self, group: str):
        """Получить старые данные"""
        cache_key = f"data_{group}"
        try:
            if os.path.exists(self.cache_file):
                with open(self.cache_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    return data.get(cache_key)
        except Exception as e:
            print(f"❌ Ошибка чтения старых данных: {e}")
        return None

    def save_schedule_data(self, group: str, schedule_data: dict, hash_value: str):
        """Сохранить данные расписания"""
        cache_key = f"data_{group}"
        try:
            cache_data = {}
            if os.path.exists(self.cache_file):
                with open(self.cache_file, 'r', encoding='utf-8') as f:
                    cache_data = json.load(f)
            
            cache_data[cache_key] = {
                'data': schedule_data,
                'hash': hash_value,
                'last_update': datetime.now().isoformat(),
                'last_update_human': datetime.now().strftime("%d.%m.%Y %H:%M:%S")
            }
            
            with open(self.cache_file, 'w', encoding='utf-8') as f:
                json.dump(cache_data, f, ensure_ascii=False, indent=2)
            
            print(f"💾 Данные сохранены для {group}")
            
        except Exception as e:
            print(f"❌ Ошибка сохранения данных: {e}")

    def force_update_cache(self, group: str):
        """Принудительное обновление кэша для группы"""
        print(f"💾 Принудительное обновление кэша для группы {group}")
        
        current_data = self.extract_schedule_data(group)
        if not current_data:
            print(f"❌ Не удалось извлечь данные для группы {group}")
            return False
        
        current_hash = self.calculate_smart_hash(group)
        if not current_hash:
            print(f"❌ Не удалось вычислить хэш для группы {group}")
            return False
        
        self.save_schedule_data(group, current_data, current_hash)
        print(f"✅ Кэш обновлен для группы {group}")
        return True

    def clear_cache(self):
        """Очистка всего кэша"""
        try:
            if os.path.exists(self.cache_file):
                os.remove(self.cache_file)
                print("🗑️ Кэш полностью очищен")
                return True
        except Exception as e:
            print(f"❌ Ошибка очистки кэша: {e}")
        return False

    def debug_group(self, group: str):
        """Отладочная информация по группе"""
        print(f"🔍 ОТЛАДКА ГРУППЫ: {group}")
        
        current_data = self.extract_schedule_data(group)
        if not current_data:
            print("❌ Не удалось извлечь текущие данные")
            return
        
        current_hash = self.calculate_smart_hash(group)
        old_data = self.get_old_data(group)
        
        print(f"📊 Текущий хэш: {current_hash}")
        print(f"📊 Старый хэш: {old_data.get('hash', 'Нет данных') if old_data else 'Нет данных'}")
        print(f"📊 Изменения: {current_hash != old_data.get('hash') if old_data else 'Нет сравнения'}")
        
        # Показать текущие данные
        print("\n📋 ТЕКУЩИЕ ДАННЫЕ:")
        for week_type in ['even', 'odd']:
            if week_type in current_data.get('weeks', {}):
                print(f"  📅 {week_type.upper()} НЕДЕЛЯ:")
                for day, day_data in current_data['weeks'][week_type].items():
                    print(f"    📝 {day}:")
                    schedule_data = day_data.get('schedule', [])
                    print(f"      schedule ({len(schedule_data)} строк):")
                    for i, row in enumerate(schedule_data):
                        if any(cell.strip() for cell in row):  # Показываем только непустые строки
                            print(f"        Строка {i+1}: {row}")

    def get_cache_info_for_group(self, group: str):
        """Получить информацию о кэше для группы"""
        cache_key = f"data_{group}"
        try:
            if os.path.exists(self.cache_file):
                with open(self.cache_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    group_data = data.get(cache_key, {})
                    hash_value = group_data.get('hash', 'Неизвестно')
                    return {
                        'last_checked': group_data.get('last_update_human', 'Неизвестно'),
                        'hash': hash_value[:16] + '...' if hash_value != 'Неизвестно' else 'Неизвестно',
                        'group': group
                    }
        except Exception as e:
            print(f"❌ Ошибка чтения информации о кэше для группы {group}: {e}")
        
        return {'last_checked': 'Неизвестно', 'hash': 'Неизвестно', 'group': group}

    def force_detect_changes(self, group: str = None):
        """Принудительная детекция изменений"""
        print("🔄 ПРИНУДИТЕЛЬНАЯ ДЕТЕКЦИЯ ИЗМЕНЕНИЙ")
        
        if group:
            groups = [group]
        else:
            groups = list(RANGES.keys())
        
        changes_found = False
        for grp in groups:
            print(f"\n🔍 Принудительная проверка группы {grp}")
            has_changed, changes = self.has_changed(grp)
            
            if has_changed:
                print(f"🎉 ИЗМЕНЕНИЯ ОБНАРУЖЕНЫ ДЛЯ {grp}!")
                changes_found = True
            else:
                print(f"✅ Изменений нет для {grp}")
        
        return changes_found

    def check_all_groups(self):
        """Проверить все группы без отправки уведомлений"""
        print("🎯 ПРОВЕРКА ВСЕХ ГРУПП")
        
        available_groups = list(RANGES.keys())
        print(f"Доступные группы: {available_groups}")
        
        changed_groups = []
        
        for group in available_groups:
            print(f"\n🔍 Проверка группы: {group}")
            has_changed, changes = self.has_changed(group)
            
            if has_changed:
                print(f"🎉 ОБНАРУЖЕНЫ ИЗМЕНЕНИЯ ДЛЯ ГРУППЫ {group}!")
                changed_groups.append((group, changes))
            else:
                print(f"✅ Изменений нет для группы {group}")
        
        return changed_groups

if __name__ == "__main__":
    # Тестирование умного детектора
    print("=== Тестирование SmartChangeDetector ===")
    
    detector = SmartChangeDetector()
    
    # Принудительная проверка всех групп
    detector.force_detect_changes()
    
    # Показать информацию о кэше
    print("\n📊 ИНФОРМАЦИЯ О КЭШЕ:")
    for group in RANGES.keys():
        cache_info = detector.get_cache_info_for_group(group)
        print(f"  {group}: последняя проверка {cache_info['last_checked']}, хэш {cache_info['hash']}")