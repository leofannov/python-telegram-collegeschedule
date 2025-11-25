import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import os
import json
from config import RANGES, WEEK_CONFIG, EXCEL_FILE, LAST_UPDATE_FILE

class ScheduleParser:
    def __init__(self):
        self.ranges = RANGES
        self.week_config = WEEK_CONFIG
        self._cache = {}
        self._cache_file = 'cache/schedule_data.cache'
        self._load_cache()

    def _load_cache(self):
        """Загрузка кэша из файла"""
        try:
            if os.path.exists(self._cache_file):
                with open(self._cache_file, 'r', encoding='utf-8') as f:
                    self._cache = json.load(f)
                print(f"✅ Загружен кэш из файла: {len(self._cache)} записей")
        except Exception as e:
            print(f"❌ Ошибка загрузки кэша: {e}")
            self._cache = {}

    def _save_cache(self):
        """Сохранение кэша в файл"""
        try:
            with open(self._cache_file, 'w', encoding='utf-8') as f:
                json.dump(self._cache, f, ensure_ascii=False, indent=2)
            print(f"💾 Кэш сохранен в файл: {len(self._cache)} записей")
        except Exception as e:
            print(f"❌ Ошибка сохранения кэша: {e}")

    def get_week_type(self):
        """Определение типа текущей недели"""
        return self.get_week_type_for_date(datetime.now())

    def get_week_type_for_date(self, date):
        """Определение типа недели для конкретной даты"""
        try:
            target_week = date.isocalendar()[1]
            base_week_type = self.week_config['base_week_type']
            base_week_number = self.week_config['base_week_number']
            
            weeks_difference = target_week - base_week_number
            
            if base_week_type == 'even':
                return 'even' if weeks_difference % 2 == 0 else 'odd'
            else:
                return 'odd' if weeks_difference % 2 == 0 else 'even'
        except:
            return 'even' if date.isocalendar()[1] % 2 == 0 else 'odd'

    def load_workbook(self):
        """Загрузка Excel файла"""
        if not os.path.exists(EXCEL_FILE):
            raise Exception('Файл расписания не найден')
        return load_workbook(EXCEL_FILE, data_only=True)

    def get_cell_range(self, worksheet, cell_range):
        """Получение значений из диапазона ячеек"""
        cells = worksheet[cell_range]
        values = []
        for row in cells:
            row_values = []
            for cell in row:
                value = str(cell.value).strip() if cell.value is not None else ""
                row_values.append(value)
            values.append(row_values)
        return values

    def get_day_schedule(self, group, week_type, day):
        """Получение расписания для конкретной группы, недели и дня"""
        if group not in self.ranges:
            raise ValueError(f"Группа {group} не найдена в конфигурации")
            
        cache_key = f"{group}_{week_type}_{day}"
        
        # Проверяем кэш
        if cache_key in self._cache:
            return self._cache[cache_key]
            
        try:
            wb = self.load_workbook()
            ws = wb.active
            
            day_ranges = self.ranges[group][week_type][day]
            
            pair_numbers = self.get_cell_range(ws, day_ranges['pair_numbers'])
            time_data = self.get_cell_range(ws, day_ranges['time'])
            schedule_data = self.get_cell_range(ws, day_ranges['schedule'])
            
            lessons = []
            for i in range(len(schedule_data)):
                pair = pair_numbers[i][0] if i < len(pair_numbers) and pair_numbers[i][0] else ""
                time_val = time_data[i][0] if i < len(time_data) and time_data[i][0] else ""
                
                row = schedule_data[i]
                content = []
                for cell in row:
                    if cell and str(cell).strip() != '':
                        content.append(str(cell).strip())
                
                if content:
                    discipline = '\n'.join(content)
                    lessons.append({
                        'pair': pair,
                        'time': time_val,
                        'discipline': discipline
                    })
            
            wb.close()
            
            # Сохраняем в кэш и в файл
            self._cache[cache_key] = lessons
            self._save_cache()
            
            return lessons
            
        except Exception as e:
            print(f"Ошибка при получении расписания для группы {group}: {e}")
            return []

    def get_week_schedule(self, group, week_type):
        """Получение расписания на всю неделю (оптимизированная версия)"""
        if group not in self.ranges:
            raise ValueError(f"Группа {group} не найдена в конфигурации")
            
        cache_key = f"{group}_{week_type}_full"
        
        if cache_key in self._cache:
            return self._cache[cache_key]
            
        try:
            wb = self.load_workbook()
            ws = wb.active
            
            week_schedule = {}
            days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']
            
            for day in days:
                day_ranges = self.ranges[group][week_type][day]
                
                pair_numbers = self.get_cell_range(ws, day_ranges['pair_numbers'])
                time_data = self.get_cell_range(ws, day_ranges['time'])
                schedule_data = self.get_cell_range(ws, day_ranges['schedule'])
                
                lessons = []
                for i in range(len(schedule_data)):
                    pair = pair_numbers[i][0] if i < len(pair_numbers) and pair_numbers[i][0] else ""
                    time_val = time_data[i][0] if i < len(time_data) and time_data[i][0] else ""
                    
                    row = schedule_data[i]
                    content = []
                    for cell in row:
                        if cell and str(cell).strip() != '':
                            content.append(str(cell).strip())
                    
                    if content:
                        discipline = '\n'.join(content)
                        lessons.append({
                            'pair': pair,
                            'time': time_val,
                            'discipline': discipline
                        })
                
                week_schedule[day] = lessons
            
            wb.close()
            
            # Сохраняем в кэш и в файл
            self._cache[cache_key] = week_schedule
            self._save_cache()
            
            return week_schedule
            
        except Exception as e:
            print(f"Ошибка при получении расписания на неделю для группы {group}: {e}")
            return {}

    def format_schedule_text(self, group, week_type, day, lessons):
        """Форматирование расписания в текст для Telegram"""
        week_type_text = "чётная" if week_type == 'even' else "нечётная"
        
        if not lessons:
            return f"📅 {day} ({week_type_text} неделя) - {group}\n\nЗанятий нет"
        
        text = f"📅 {day} ({week_type_text} неделя) - {group}\n\n"
        
        for lesson in lessons:
            if lesson['pair']:
                text += f"🔹 {lesson['pair']} пара"
                if lesson['time']:
                    text += f" ({lesson['time']})"
                text += f"\n"
            text += f"{lesson['discipline']}\n\n"
        
        return text

    def format_week_schedule_text(self, group, week_type, week_schedule):
        """Форматирование расписания на неделю"""
        week_type_text = "чётная" if week_type == 'even' else "нечётная"
        
        text = f"📅 Расписание на {week_type_text} неделю - {group}\n\n"
        
        for day, lessons in week_schedule.items():
            text += f"*{day}:*\n"
            if not lessons:
                text += "Занятий нет\n\n"
            else:
                for lesson in lessons:
                    if lesson['pair']:
                        text += f"🔹 {lesson['pair']} пара"
                        if lesson['time']:
                            text += f" ({lesson['time']})"
                        text += f"\n"
                    text += f"{lesson['discipline']}\n"
                text += "\n"
            text += "─" * 30 + "\n\n"
        
        return text

    def get_last_update(self):
        """Получение времени последнего обновления"""
        if os.path.exists(LAST_UPDATE_FILE):
            with open(LAST_UPDATE_FILE, 'r', encoding='utf-8') as f:
                return f.read().strip()
        return "Неизвестно"

    def clear_cache(self):
        """Очистка кэша в памяти и файлового кэша"""
        self._cache = {}
        try:
            if os.path.exists(self._cache_file):
                os.remove(self._cache_file)
                print("🗑️ Файловый кэш очищен")
        except Exception as e:
            print(f"❌ Ошибка очистки файлового кэша: {e}")