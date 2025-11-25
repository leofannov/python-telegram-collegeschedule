import requests
import os
import sys
import hashlib
from datetime import datetime
from config import EXCEL_FILE, LAST_UPDATE_FILE, TOKEN
from change_notifier import ChangeNotifier
import tempfile
import shutil
import logging
from openpyxl import load_workbook

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('cache/download.log', encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)

def download_schedule():
    """Скачивание расписания и проверка изменений"""
    url = ""
    
    try:
        print("=" * 60)
        print("🎯 НАЧАЛО ПРОВЕРКИ РАСПИСАНИЯ")
        print("=" * 60)
        
        # Создаем папку cache если нет
        if not os.path.exists('cache'):
            os.makedirs('cache')
            print("📁 Создана папка cache")

        # ШАГ 1: Скачиваем новое расписание
        print(f"📥 Скачивание нового расписания с {url}...")
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        
        # Проверяем размер файла
        file_size = len(response.content)
        if file_size < 1024:  # Меньше 1KB
            print(f"❌ Скачанный файл слишком маленький: {file_size} байт")
            return False
        
        # Создаем временный файл для проверки
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            temp_path = temp_file.name
            temp_file.write(response.content)
        
        # ШАГ 2: Проверяем что файл валиден
        print("🔍 Проверка валидности Excel файла...")
        try:
            wb = load_workbook(temp_path)
            sheet = wb.active
            
            # Базовая проверка что файл содержит данные
            if sheet.max_row < 10 or sheet.max_column < 5:
                print("❌ Файл не содержит достаточное количество данных")
                wb.close()
                os.unlink(temp_path)
                return False
                
            wb.close()
            print("✅ Скачанный файл валиден")
            
        except Exception as e:
            print(f"❌ Скачанный файл поврежден или не является Excel файлом: {e}")
            os.unlink(temp_path)
            return False
        
        # ШАГ 3: Заменяем старый файл на новый (ВСЕГДА)
        if os.path.exists(EXCEL_FILE):
            # Создаем резервную копию
            backup_path = EXCEL_FILE + '.backup'
            shutil.copy2(EXCEL_FILE, backup_path)
            print(f"📦 Создана резервная копия: {backup_path}")
        
        shutil.move(temp_path, EXCEL_FILE)
        
        # Обновляем время последнего обновления
        update_time = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
        with open(LAST_UPDATE_FILE, 'w', encoding='utf-8') as f:
            f.write(update_time)
        
        print(f"✅ Расписание успешно обновлено: {update_time}")
        print(f"📊 Размер файла: {file_size} байт")
        
        # ШАГ 4: Проверяем изменения и отправляем уведомления (ПЕРЕД очисткой кэша)
        print("\n🔍 ПРОВЕРКА ИЗМЕНЕНИЙ ДЛЯ ВСЕХ ГРУПП...")
        notifier = ChangeNotifier()
        
        # Принудительно проверяем изменения (новый файл против старого кэша)
        changes_detected = notifier.check_changes_after_download(TOKEN)
        
        if changes_detected:
            print("🎉 УВЕДОМЛЕНИЯ ОБ ИЗМЕНЕНИЯХ ОТПРАВЛЕНЫ")
        else:
            print("ℹ️ Изменений не обнаружено")
        
        # ШАГ 5: Очищаем кэш парсера ПОСЛЕ проверки изменений
        print("\n🗑️ Очистка кэша парсера для обновления данных бота...")
        try:
            from schedule_parser import ScheduleParser
            parser = ScheduleParser()
            parser.clear_cache()
            print("✅ Очищен кэш парсера - бот будет использовать обновленные данные")
        except Exception as e:
            print(f"⚠️ Не удалось очистить кэш парсера: {e}")
        
        # ШАГ 6: Устанавливаем флаг перезагрузки для бота
        print("\n🔍 ПРОВЕРКА ИЗМЕНЕНИЙ ДЛЯ ВСЕХ ГРУПП...")
        notifier = ChangeNotifier()
        changes_detected = notifier.check_changes_after_download(TOKEN)

        if changes_detected:
            print("🎉 УВЕДОМЛЕНИЯ ОБ ИЗМЕНЕНИЯХ ОТПРАВЛЕНЫ")
        else:
            print("ℹ️ Изменений не обнаружено")
            
        print("\n🔄 Установка флага перезагрузки кэша для бота...")
        try:
            reload_flag = 'cache/reload_cache.flag'
            with open(reload_flag, 'w') as f:
                f.write(datetime.now().isoformat())
            print("✅ Флаг перезагрузки кэша установлен - бот обновит данные при следующем запросе")
        except Exception as e:
            print(f"⚠️ Не удалось установить флаг перезагрузки: {e}")
        
        return True
        
    except requests.RequestException as e:
        print(f"❌ Ошибка сети при скачивании расписания: {e}")
        return False
    except Exception as e:
        print(f"❌ КРИТИЧЕСКАЯ ОШИБКА: {e}")
        import traceback
        traceback.print_exc()
        return False

def check_changes_only():
    """Только проверка изменений без скачивания"""
    try:
        print("🎯 ПРОВЕРКА ИЗМЕНЕНИЙ БЕЗ СКАЧИВАНИЯ")
        
        if not os.path.exists(EXCEL_FILE):
            print("❌ Файл расписания не найден")
            return False
            
        notifier = ChangeNotifier()
        changes_detected = notifier.check_and_notify(TOKEN)
        
        if changes_detected:
            print("🎉 Уведомления об изменениях отправлены")
        else:
            print("ℹ️ Изменений не обнаружено")
            
        return changes_detected
    except Exception as e:
        print(f"❌ Ошибка при проверке изменений: {e}")
        import traceback
        traceback.print_exc()
        return False

def force_check_changes():
    """Принудительная проверка изменений"""
    try:
        print("🎯 ПРИНУДИТЕЛЬНАЯ ПРОВЕРКА ИЗМЕНЕНИЙ")
        notifier = ChangeNotifier()
        changes_detected = notifier.force_check_and_notify(TOKEN)
        
        if changes_detected:
            print("🎉 Уведомления об изменениях отправлены")
        else:
            print("ℹ️ Изменений не обнаружено")
            
        return changes_detected
    except Exception as e:
        print(f"❌ Ошибка при проверке изменений: {e}")
        import traceback
        traceback.print_exc()
        return False

def show_statistics():
    """Показать статистику уведомлений"""
    try:
        print("=== Статистика уведомлений ===")
        notifier = ChangeNotifier()
        stats = notifier.get_statistics()
        
        print(f"Всего чатов в системе: {stats['total_chats']}")
        print(f"Чатов с уведомлениями: {stats['enabled_chats']}")
        print(f"Чатов без уведомлений: {stats['disabled_chats']}")
        
        # Показать список включенных чатов
        enabled_chats = notifier.get_all_enabled_chats()
        print(f"\nВключенные чаты ({len(enabled_chats)}):")
        for chat_id in enabled_chats:
            group = notifier.group_manager.get_group(chat_id)
            print(f"  - {chat_id} (группа: {group})")
        
        return stats
    except Exception as e:
        print(f"❌ Ошибка при получении статистики: {e}")
        import traceback
        traceback.print_exc()
        return None

def test_notification():
    """Тестовая отправка уведомления"""
    try:
        print("=== Тестовая отправка уведомления ===")
        notifier = ChangeNotifier()
        return notifier.send_test_notification(TOKEN)
    except Exception as e:
        print(f"❌ Ошибка при тестовой отправке: {e}")
        import traceback
        traceback.print_exc()
        return False

def force_detect_changes():
    """Принудительная детекция изменений для всех групп"""
    try:
        print("🎯 ПРИНУДИТЕЛЬНАЯ ДЕТЕКЦИЯ ИЗМЕНЕНИЙ ДЛЯ ВСЕХ ГРУПП")
        notifier = ChangeNotifier()
        return notifier.force_detect_changes()
    except Exception as e:
        print(f"❌ Ошибка при принудительной детекции: {e}")
        import traceback
        traceback.print_exc()
        return False

def debug_group(group_name):
    """Отладочная функция: показать ячейки группы"""
    try:
        print(f"🔍 ОТЛАДКА ГРУППЫ: {group_name}")
        notifier = ChangeNotifier()
        notifier.debug_group(group_name)
    except Exception as e:
        print(f"❌ Ошибка отладки: {e}")

def clear_parser_cache():
    """Очистка только кэша парсера (без кэша детектора)"""
    try:
        print("🗑️ ОЧИСТКА КЭША ПАРСЕРА")
        
        # Очищаем кэш парсера
        try:
            from schedule_parser import ScheduleParser
            parser = ScheduleParser()
            parser.clear_cache()
            print("✅ Очищен кэш парсера")
        except Exception as e:
            print(f"⚠️ Не удалось очистить кэш парсера: {e}")
        
        print("🎉 Кэш парсера очищен - бот будет перезагружать данные")
        return True
        
    except Exception as e:
        print(f"❌ Ошибка при очистке кэша парсера: {e}")
        import traceback
        traceback.print_exc()
        return False

def clear_all_caches():
    """Очистка всех кэшей (только для экстренных случаев)"""
    try:
        print("🚨 ОЧИСТКА ВСЕХ КЭШЕЙ (ЭКСТРЕННАЯ)")
        
        # Очищаем кэш парсера
        try:
            from schedule_parser import ScheduleParser
            parser = ScheduleParser()
            parser.clear_cache()
            print("✅ Очищен кэш парсера")
        except Exception as e:
            print(f"⚠️ Не удалось очистить кэш парсера: {e}")
        
        # Очищаем кэш умного детектора
        try:
            from smart_change_detector import SmartChangeDetector
            smart_detector = SmartChangeDetector()
            smart_detector.clear_cache()
            print("✅ Очищен кэш умного детектора")
        except Exception as e:
            print(f"⚠️ Не удалось очистить кэш умного детектора: {e}")
        
        # Очищаем файловые кэши
        cache_files = [
            'cache/schedule_data.cache',
            'cache/schedule_hash.cache',
            'cache/smart_schedule_cache.json'
        ]
        
        for cache_file in cache_files:
            if os.path.exists(cache_file):
                os.remove(cache_file)
                print(f"✅ Удален {cache_file}")
        
        print("🎉 Все кэши очищены - система начнет с чистого листа")
        return True
        
    except Exception as e:
        print(f"❌ Ошибка при очистке кэшей: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    # Обработка аргументов командной строки
    if len(sys.argv) > 1:
        command = sys.argv[1]
        if command == "check":
            # Только проверка изменений
            check_changes_only()
        elif command == "force_check":
            # Принудительная проверка изменений
            force_check_changes()
        elif command == "stats":
            # Показать статистику
            show_statistics()
        elif command == "test":
            # Тестовая отправка
            test_notification()
        elif command == "force_detect":
            # Принудительная детекция
            force_detect_changes()
        elif command == "debug" and len(sys.argv) > 2:
            # Отладка конкретной группы
            debug_group(sys.argv[2])
        elif command == "clear_parser_cache":
            # Очистка только кэша парсера
            clear_parser_cache()
        elif command == "clear_all_caches":
            # Очистка всех кэшей (экстренная)
            clear_all_caches()
        elif command == "help":
            print("Доступные команды:")
            print("  python cron_download.py                    - скачать расписание и проверить изменения")
            print("  python cron_download.py check              - только проверка изменений (без скачивания)")
            print("  python cron_download.py force_check        - принудительная проверка изменений")
            print("  python cron_download.py stats              - показать статистику")
            print("  python cron_download.py test               - тестовая отправка уведомления")
            print("  python cron_download.py force_detect       - принудительная детекция изменений")
            print("  python cron_download.py debug <group>      - отладка конкретной группы")
            print("  python cron_download.py clear_parser_cache - очистка только кэша парсера (без детектора)")
            print("  python cron_download.py clear_all_caches   - очистка всех кэшей (экстренная)")
            print("  python cron_download.py help               - показать эту справку")
        else:
            print(f"Неизвестная команда: {command}")
            print("Используйте 'python cron_download.py help' для справки")
    else:
        # Обычная загрузка расписания и проверка изменений
        success = download_schedule()
        sys.exit(0 if success else 1)